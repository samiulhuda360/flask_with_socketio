import os
import time
import csv
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash, url_for
from services import (get_all_sitenames, get_url_data_from_db, save_matched_to_excel, process_site, store_posted_url, extract_domain, delete_site_and_links)
from flask_socketio import SocketIO, emit
import json
from functools import wraps
import openpyxl
from utils import get_api_keys
import pandas as pd
import sqlite3
from werkzeug.utils import secure_filename


app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['SESSION_COOKIE_SAMESITE'] = "Lax"  # or "None" if necessary
app.config['SESSION_COOKIE_SECURE'] = False

app.secret_key = 'sdfadfasdfasdfasdfasdf'

uploaded_filename = None
Exact_MATCH = False
SKIP_COM_AU = False
ONLY_COM_AU = False
#DataBase Operation
def get_link_list_from_db(host_url):
    # Establish a connection to the SQLite database
    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    # Query to retrieve all the links for a given site
    cursor.execute('''
    SELECT url 
    FROM links
    INNER JOIN sites ON links.site_id = sites.site_id
    WHERE sitename = ?
    ''', (host_url,))

    links = [row[0] for row in cursor.fetchall()]
    conn.close()

    return links

@app.route('/save_api_config', methods=['POST'])
def save_api_config():
    openai_api = request.form.get('openaiapi')
    pexels_api = request.form.get('pexelsapi')

    con = sqlite3.connect('api_config.db')
    cur = con.cursor()

    # Check if the table already contains data
    cur.execute("SELECT COUNT(*) FROM api_keys")
    count = cur.fetchone()[0]

    if count == 0:
        # Insert new APIs
        cur.execute("INSERT INTO api_keys (openai_api, pexels_api) VALUES (?, ?)", (openai_api, pexels_api))
    else:
        # Update existing APIs depending on which fields are filled
        if openai_api and pexels_api:
            cur.execute("UPDATE api_keys SET openai_api = ?, pexels_api = ? WHERE id = 1", (openai_api, pexels_api))
        elif openai_api:
            cur.execute("UPDATE api_keys SET openai_api = ? WHERE id = 1", (openai_api,))
        elif pexels_api:
            cur.execute("UPDATE api_keys SET pexels_api = ? WHERE id = 1", (pexels_api,))

    con.commit()
    con.close()

    flash("API configuration updated successfully!", "success")
    return redirect(url_for('config_manager'))

# Flask route to handle site deletion
@app.route('/delete-site', methods=['DELETE'])
def delete_site_route():
    sitename = request.args.get('sitename')
    if not sitename or sitename == "all":
        return jsonify({"error": "Invalid site name"}), 400

    delete_site_and_links(sitename)

    return jsonify({"message": "Site and associated links deleted successfully"}), 200

@app.route('/get_files')
def get_files():
    folder = app.config['UPLOAD_FOLDER']
    files = os.listdir(folder)
    return jsonify(files)
@app.route('/download_excel')
def download_excel():
    global uploaded_filename
    print(uploaded_filename)

    if not uploaded_filename:
        return jsonify({"error": "No file available for download"}), 404

    excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_filename)

    # Send the file as a response with appropriate headers
    try:
        return send_file(excel_file_path, as_attachment=True, download_name=uploaded_filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/failed_csv')
def failed_site_download_excel():
    # Specify the file path where the Excel file is saved
    failed_csv_file_path = 'failed_urls.csv'  # Adjust the path as needed

    # Send the file as a response with appropriate headers
    try:
        return send_file(failed_csv_file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/uploaded_excel')
def uploaded_download_excel():
    # Specify the file path where the Excel file is saved
    uploaded_excel_file_path = 'uploads/uploaded_excel.xlsx'  # Adjust the path as needed

    # Send the file as a response with appropriate headers
    try:
        return send_file(uploaded_excel_file_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Dummy logic: replace with real authentication logic
        username = request.form['username']
        password = request.form['password']
        print(f"Received username: {username}, password: {password}")

        if username == "admin" and password == "p@ssword123":  # Dummy check, replace with real logic
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            flash('Incorrect credentials.')
            return redirect(url_for('login'))
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('logged_in', None)  # Clear the logged_in flag from the session.
    flash('You were successfully logged out.')
    return redirect(url_for('login'))  # Redirect to login page or another page of your choice.


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)

    return decorated_function


@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/config', methods=['GET', 'POST'])
@login_required
def config_manager():
    if request.method == 'POST':
        openai_api = request.form.get('openaiapi').strip()
        pexels_api = request.form.get('pexelsapi').strip()

        con = sqlite3.connect('api_config.db')
        cur = con.cursor()

        # Check if the table already contains data
        cur.execute("SELECT COUNT(*) FROM api_keys")
        count = cur.fetchone()[0]

        if count == 0:
            # Insert new APIs
            cur.execute("INSERT INTO api_keys (openai_api, pexels_api) VALUES (?, ?)", (openai_api, pexels_api))
        else:
            # Update existing APIs depending on which fields are filled
            if openai_api and pexels_api:
                cur.execute("UPDATE api_keys SET openai_api = ?, pexels_api = ? WHERE id = 1", (openai_api, pexels_api))
            elif openai_api:
                cur.execute("UPDATE api_keys SET openai_api = ? WHERE id = 1", (openai_api,))
            elif pexels_api:
                cur.execute("UPDATE api_keys SET pexels_api = ? WHERE id = 1", (pexels_api,))

        con.commit()
        con.close()

        flash("API configuration updated successfully!", "success")

    # Fetch the keys to display
    api_keys = get_api_keys() or {"openai_api": "", "pexels_api": ""}
    return render_template('configuration.html', openai_api=api_keys["openai_api"], pexels_api=api_keys["pexels_api"])


@app.route('/site-manager', methods=['GET'])
@login_required
def site_manager():
    sitename_filter = request.args.get('sitename_filter', None)  # Notice the default value is now None

    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    sites_data = []

    if sitename_filter:
        if sitename_filter == 'all':
            cursor.execute('SELECT * FROM sites')
        else:
            cursor.execute('SELECT * FROM sites WHERE sitename = ?', (sitename_filter,))

        sites = cursor.fetchall()

        for site in sites:
            site_id, sitename, username, app_password = site
            cursor.execute('SELECT url FROM links WHERE site_id = ?', (site_id,))
            links = cursor.fetchall()
            sites_data.append({
                'site_id': site_id,
                'sitename': sitename,
                'username': username,
                'app_password': app_password,
                'links': [link[0] for link in links]
            })

    # Fetching all site names for the dropdown
    cursor.execute('SELECT DISTINCT sitename FROM sites')
    all_sitenames = [row[0] for row in cursor.fetchall()]

    conn.close()

    return render_template('site_manager.html', filtered_sites=sites_data, all_sitenames=all_sitenames)





@app.route('/upload-excel', methods=['POST'])
@login_required
def upload_excel_site_data():
    if 'excel_data' not in request.files:
        return "No file part", 400

    file = request.files['excel_data']

    if file.filename == '':
        return "No selected file", 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return "Invalid file type. Please upload an Excel file.", 400

    df = pd.read_excel(file, engine='openpyxl')

    conn = sqlite3.connect('sites_data.db')
    cursor = conn.cursor()

    for _, row in df.iterrows():
        sitename, username, app_password, link = row['sitename'], row['username'], row['app_password'], row['links']

        cursor.execute('SELECT site_id FROM sites WHERE sitename=?', (sitename,))
        site_id = cursor.fetchone()

        if site_id:  # if site exists
            cursor.execute('UPDATE sites SET username=?, app_password=? WHERE site_id=?',
                           (username, app_password, site_id[0]))
        else:  # if site doesn't exist
            cursor.execute('INSERT INTO sites (sitename, username, app_password) VALUES (?, ?, ?)',
                           (sitename, username, app_password))
            site_id = (cursor.lastrowid,)

        # Insert link if it doesn't exist for the site
        cursor.execute('SELECT url FROM links WHERE site_id=? AND url=?', (site_id[0], link))
        if not cursor.fetchone():
            cursor.execute('INSERT INTO links (site_id, url) VALUES (?, ?)', (site_id[0], link))

    conn.commit()
    conn.close()
    return redirect(url_for('site_manager'))



def update_excel_with_live_link(file_path, row_index, live_url):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Assuming you want to write to column 'J=Live_link'
    sheet[f'J{row_index}'] = live_url

    wb.save(file_path)

@app.route('/stop_processing', methods=['POST'])
def stop_processing():
    global should_continue_processing
    should_continue_processing = False
    return jsonify({"message": "Processing will be stopped."}), 200


@app.route('/start_emit', methods=['POST'])
def start_emit():
    global should_continue_processing
    global USE_IMAGES
    should_continue_processing = True

    # Get the Excel file
    excel_file = request.files['excel_file']
    if not excel_file:
        return jsonify({"error": "No file uploaded."}), 400

    # Set USE_IMAGES based on checkbox value
    USE_IMAGES = 'use_images' in request.form
    Exact_MATCH = 'exact_match' in request.form
    SKIP_COM_AU = 'skip_au' in request.form
    ONLY_COM_AU = 'only_au' in request.form

    global uploaded_filename
    original_filename = secure_filename(excel_file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], original_filename)
    excel_file.save(file_path)

    # Store the filename in the global variable
    uploaded_filename = original_filename
    print(uploaded_filename)

    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Create an empty list to accumulate the data
    data_list = []
    sitenames = get_all_sitenames()
    num_sites = len(sitenames)

    total_rows = len(list(sheet.iter_rows(values_only=True)))
    row_index = 0

    last_used_site_index = -1  # Start with -1 so that for the first row, it starts with 0.
    with open('failed_urls.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Failed URLs"])
    try:
        while row_index < total_rows and should_continue_processing:
            row = list(sheet.iter_rows(values_only=True))[row_index]

            if row_index == 0:
                row_index += 1
                continue
            if row[1] is None or row[2] is None:
                row_index += 1
                continue

                # asyncio.run(my_async_function())
            anchor, linking_url, embed_code, map_embed_title, name, address, phone, topic, live_link = row[1:10]


            if row[9] != None and row[9] != "Failed To Post":
                print("skipping the row")
                row_index += 1
                continue


            link_posted = False
            failed_post_count = 0  # Counter for failed posts.
            start_site_index = (last_used_site_index + 1) % num_sites  # Start from the next site.

            for offset in range(num_sites):
                site_index = (start_site_index + offset) % num_sites  # Wrap around using modulo.
                host_url = sitenames[site_index].strip()
                print("Posting to:", host_url)
                link_list = get_link_list_from_db(host_url)
                # Extract domains from link_list
                domain_list = [extract_domain(link) for link in link_list]

                # Inside the success condition where you've successfully posted the link:
                last_used_site_index = site_index  # Update the last used site index.

                # Exact Match Check
                if Exact_MATCH:
                    if linking_url in link_list:
                        print(linking_url)
                        continue  # Skip to the next iteration

                # If Exact_MATCH is False, then it's Root Match
                elif extract_domain(linking_url) in domain_list:
                    print(linking_url)
                    print("Matched Root Domain inside")
                    continue  # Skip to the next iteration
                # Skip com.au and org.au Check
                if SKIP_COM_AU:
                    print("Checking URL:", host_url)
                    if host_url.endswith('com.au') or host_url.endswith('org.au'):
                        print("Skipping because it ends with com.au or org.au")
                        continue  # Skip to the next iteration

                # Only com.au and org.au Check
                if ONLY_COM_AU:
                    if not (host_url.endswith('com.au') or host_url.endswith('org.au')):
                        print("Skipping because it doesn't end with com.au or org.au")
                        continue  # Skip to the next iteration



                if not should_continue_processing:
                    break


                user_password_data = get_url_data_from_db(host_url)
                site_json = "https://" + host_url + "/wp-json/wp/v2"

                if user_password_data:
                    user = user_password_data.get('user')
                    password = user_password_data.get('password')

                    # for _ in range(5):
                    #     asyncio.run(my_async_function())
                    if not should_continue_processing:
                        break

                    # Testing START
                    nap = name + "<br>" + address + "<br>" + phone +"<br>"
                    live_url = process_site(site_json, host_url, user, password, topic, anchor, linking_url, embed_code,
                                            map_embed_title, nap, USE_IMAGES)

                    if live_url == "Failed To Post":
                        with open('failed_urls.csv', 'a', newline='') as f:
                            writer = csv.writer(f)
                            writer.writerow([host_url])
                            print("Failed at:", host_url)
                        failed_post_count += 1
                        if failed_post_count < 5:
                            continue
                        else:
                            break
                    update_excel_with_live_link(file_path, row_index + 1, live_url) # Updating Excel

                    if live_url != "Failed To Post":
                        # Adding to Database
                        store_posted_url(host_url, linking_url) # Adding the posted link to database
                    else:
                        print("Not Posting")
                    # Testing END

                    data = {
                        'id': row_index,
                        'anchor': anchor,
                        'linking_url': linking_url,
                        'nap': nap,
                        'topic': topic,
                        'live_url': live_url,
                        "Host_site": host_url,
                    }
                    time.sleep(1)

                    data_list.append(data)
                    socketio.emit('update', {'data': json.dumps(data_list)})


                    link_posted = True
                    break

            if not link_posted:
                print("All Sites have this link")
                pass

            row_index += 1

        if not should_continue_processing:
            print("Process stopped")
            flash("Process stopped")
            socketio.emit('update', {"message": "Process stopped"})
            return jsonify({"message": "Processing was halted by the user."}), 200
    except Exception as e:
        # Log the error for debugging purposes
        print(f"An error occurred: {str(e)}")
        # Send an error message to the frontend
        socketio.emit('error', {'message': str(e)})

    flash("Processed successfully!")
    socketio.emit('update', {"message": "Processing Ended"})
    return jsonify({"message": "Processing Ended"}), 200


if __name__ == '__main__':
    socketio.run(app, debug=True)

